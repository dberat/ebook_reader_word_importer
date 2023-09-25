import sqlite3
import string
from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import Entry


def process_words():

    book_name = book_name_entry.get()
    sheet_name = sheet_name_entry.get()

    # Checking if the 'words.xlsx' file already exists
    try:
        wb = load_workbook("words.xlsx")
    except FileNotFoundError:
        # If it doesn't exist, create a new workbook
        wb = Workbook()

    # Connecting to the database
    conn = sqlite3.connect('KoboReader.sqlite')
    cur = conn.cursor()

    # Executing the SQL query to fetch the Text values from the WordList table
    cur.execute(f"SELECT Text FROM WordList WHERE VolumeId = '{book_name}' ")

    # Getting all the words returned by the query
    rows = cur.fetchall()

    # Processing the words as needed
    rows = [(s[0].strip(string.punctuation),) if s[0][-1] in string.punctuation else s for s in rows]
    add_comma = lambda x: (x[0]+',',)
    rows = list(map(add_comma, rows))
    rows = list(set(rows))

    # Creating a new Excel workbook object
    wb = Workbook()
    ws = wb.create_sheet(f"{sheet_name}")

    # Writing the data from the Text column to the sheet
    for row in rows:
        ws.append(row)

    # Saving the workbook to a file
    wb.save("words.xlsx")

    cur.close()
    conn.close()

# Creating the main window
root = tk.Tk()
root.title("SQLite to Excel Converter")


# Setting the dimensions of the main window (width x height)
root.geometry("400x200")

label = tk.Label(root, text="Enter the book name:")
label.pack()

book_name_entry = Entry(root)
book_name_entry.pack()

label = tk.Label(root, text="Enter the sheet name you want to create:")
label.pack()

sheet_name_entry = Entry(root)
sheet_name_entry.pack()


process_button = tk.Button(root, text="Process Data", command=process_words)
process_button.pack()

root.mainloop()
